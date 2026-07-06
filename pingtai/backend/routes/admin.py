"""管理员路由"""
from flask import Blueprint, request, jsonify, current_app, send_from_directory, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import User, Facility, BookingRule, Booking, Feedback, OperationLog, Review, UserBehavior, Favorite, FacilityView
from utils.decorators import role_required, log_operation
from extensions import db
from datetime import datetime, timedelta, time
from sqlalchemy import func
import json
import os
import uuid

admin_bp = Blueprint('admin', __name__)

def parse_time(time_str):
    """将字符串转换为time对象"""
    if isinstance(time_str, time):
        return time_str
    if not time_str:
        return time(8, 0)
    try:
        return datetime.strptime(time_str, '%H:%M').time()
    except:
        return time(8, 0)

# ==================== 文件上传 ====================

@admin_bp.route('/upload/image', methods=['POST'])
@jwt_required()
@role_required('admin')
def upload_image():
    """上传设施图片"""
    try:
        print(f"[DEBUG] upload_image 被调用")
        print(f"[DEBUG] 当前用户: {get_jwt_identity()}")
        
        if 'file' not in request.files:
            print(f"[DEBUG] 没有找到文件")
            return jsonify({'code': 400, 'message': '没有上传文件'}), 400
        
        file = request.files['file']
        print(f"[DEBUG] 文件名: {file.filename}")
        
        if file.filename == '':
            return jsonify({'code': 400, 'message': '文件名为空'}), 400
        
        # 检查文件类型
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        def allowed_file(filename):
            return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
        
        if not allowed_file(file.filename):
            print(f"[DEBUG] 不支持的文件类型: {file.filename}")
            return jsonify({'code': 400, 'message': '不支持的文件类型'}), 400
        
        # 创建上传目录
        upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'facilities')
        print(f"[DEBUG] 上传目录: {upload_folder}")
        os.makedirs(upload_folder, exist_ok=True)
        
        # 生成唯一文件名
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(upload_folder, filename)
        
        # 保存文件
        file.save(filepath)
        print(f"[DEBUG] 文件保存成功: {filepath}")
        
        # 自动从请求中获取服务器地址，避免手动配置IP
        # 从请求头获取 host，格式可能是 "IP:端口" 或 "域名"
        host = request.host
        # 判断是否使用 https
        protocol = 'https' if request.is_secure else 'http'
        image_url = f"{protocol}://{host}/static/uploads/facilities/{filename}"
        
        print(f"[DEBUG] 返回URL: {image_url}")
        
        return jsonify({
            'code': 200,
            'message': '上传成功',
            'data': {
                'url': image_url,
                'filename': filename
            }
        })
    except Exception as e:
        import traceback
        print(f"上传失败: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'code': 500, 'message': f'上传失败: {str(e)}'}), 500

# 注意: 静态文件由Flask自动提供，无需额外路由

# ==================== 用户管理 ====================

@admin_bp.route('/users/list', methods=['GET'])
@jwt_required()
@role_required('admin')
def list_users():
    """获取用户列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    role = request.args.get('role', None)
    
    print(f'[Admin] 获取用户列表 - role: {role}')  # 调试日志
    
    query = User.query
    
    if role and role != 'all':
        query = query.filter_by(role=role)
    
    query = query.order_by(User.created_at.desc())
    pagination = query.paginate(page=page, per_page=page_size, error_out=False)
    
    print(f'[Admin] 找到 {pagination.total} 个用户')  # 调试日志
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'total': pagination.total,
            'page': page,
            'page_size': page_size,
            'users': [u.to_dict() for u in pagination.items]
        }
    })

@admin_bp.route('/users/create', methods=['POST'])
@jwt_required()
@role_required('admin')
@log_operation('create_user', 'user_management')
def create_user():
    """创建用户（普通用户或审核员）"""
    data = request.get_json()
    
    required_fields = ['phone', 'password', 'username', 'role']
    for field in required_fields:
        if field not in data:
            return jsonify({'code': 400, 'message': f'缺少必填字段: {field}'}), 400
    
    phone = data['phone']
    role = data['role']
    
    # 验证角色
    if role not in ['resident', 'auditor']:
        return jsonify({'code': 400, 'message': '角色只能是居民或审核员'}), 400
    
    # 验证手机号格式
    import re
    if not re.match(r'^1\d{10}$', phone):
        return jsonify({'code': 400, 'message': '手机号格式不正确，必须是11位数字且以1开头'}), 400
    
    # 检查手机号是否已存在
    if User.query.filter_by(phone=phone).first():
        return jsonify({'code': 400, 'message': '该手机号已注册'}), 400
    
    # 创建用户
    user = User(
        phone=phone,
        username=data['username'],
        role=role
    )
    user.set_password(data['password'])
    
    try:
        db.session.add(user)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': user.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] 创建用户失败: {str(e)}")
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500

@admin_bp.route('/users/create-auditor', methods=['POST'])
@jwt_required()
@role_required('admin')
@log_operation('create_auditor', 'user_management')
def create_auditor():
    """创建审核员账户（兼容旧接口）"""
    data = request.get_json()
    
    required_fields = ['phone', 'password', 'username']
    for field in required_fields:
        if field not in data:
            return jsonify({'code': 400, 'message': f'缺少必填字段: {field}'}), 400
    
    phone = data['phone']
    
    # 验证手机号格式
    import re
    if not re.match(r'^1\d{10}$', phone):
        return jsonify({'code': 400, 'message': '手机号格式不正确，必须是11位数字且以1开头'}), 400
    
    # 检查手机号是否已存在
    if User.query.filter_by(phone=phone).first():
        return jsonify({'code': 400, 'message': '该手机号已注册'}), 400
    
    # 创建审核员
    user = User(
        phone=phone,
        username=data['username'],
        role='auditor'
    )
    user.set_password(data['password'])
    
    try:
        db.session.add(user)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': user.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] 创建审核员失败: {str(e)}")
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500

@admin_bp.route('/users/toggle-status/<int:user_id>', methods=['POST'])
@jwt_required()
@role_required('admin')
@log_operation('toggle_user_status', 'user_management')
def toggle_user_status(user_id):
    """启用/禁用用户"""
    user = db.session.get(User, user_id)
    
    if not user:
        return jsonify({'code': 404, 'message': '用户不存在'}), 404
    
    # 不能禁用自己
    current_user_id = get_jwt_identity()
    if user_id == current_user_id:
        return jsonify({'code': 400, 'message': '不能禁用自己'}), 400
    
    user.status = 0 if user.status == 1 else 1
    
    try:
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '操作成功',
            'data': user.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'操作失败: {str(e)}'}), 500


@admin_bp.route('/users/delete/<int:user_id>', methods=['DELETE'])
@jwt_required()
@role_required('admin')
@log_operation('delete_user', 'user_management')
def delete_user(user_id):
    """删除用户"""
    from models import Booking, Notification, UserBehavior, FacilityView, Favorite, CheckinRecord, Review

    user = db.session.get(User, user_id)

    if not user:
        return jsonify({'code': 404, 'message': '用户不存在'}), 404

    # 不能删除自己
    current_user_id = get_jwt_identity()
    if user_id == current_user_id:
        return jsonify({'code': 400, 'message': '不能删除自己'}), 400

    try:
        # 先删除关联数据
        Booking.query.filter_by(user_id=user_id).delete()
        Notification.query.filter_by(user_id=user_id).delete()
        UserBehavior.query.filter_by(user_id=user_id).delete()
        FacilityView.query.filter_by(user_id=user_id).delete()
        Favorite.query.filter_by(user_id=user_id).delete()
        CheckinRecord.query.filter_by(user_id=user_id).delete()
        Review.query.filter_by(user_id=user_id).delete()

        # 最后删除用户
        db.session.delete(user)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '删除成功'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除失败: {str(e)}'}), 500


# ==================== 设施管理 ====================

@admin_bp.route('/facilities/list', methods=['GET'])
@jwt_required()
@role_required('admin')
def list_facilities():
    """获取设施列表（管理员）"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    category = request.args.get('category', '')
    keyword = request.args.get('keyword', '')
    
    query = Facility.query
    
    # 按分类筛选
    if category:
        query = query.filter_by(category=category)
    
    # 按关键词搜索
    if keyword:
        query = query.filter(Facility.name.like(f'%{keyword}%'))
    
    query = query.order_by(Facility.created_at.desc())
    pagination = query.paginate(page=page, per_page=page_size, error_out=False)
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'total': pagination.total,
            'page': page,
            'page_size': page_size,
            'facilities': [f.to_dict() for f in pagination.items]
        }
    })

@admin_bp.route('/facilities/categories', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_facility_categories():
    """获取所有设施分类"""
    categories = db.session.query(
        Facility.category,
        func.count(Facility.facility_id).label('count')
    ).group_by(Facility.category).all()
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': [{'category': c[0], 'count': c[1]} for c in categories]
    })

@admin_bp.route('/facilities/create', methods=['POST'])
@jwt_required()
@role_required('admin')
@log_operation('create_facility', 'facility_management')
def create_facility():
    """创建设施"""
    data = request.get_json()

    required_fields = ['name', 'category', 'location']
    for field in required_fields:
        if field not in data:
            return jsonify({'code': 400, 'message': f'缺少必填字段: {field}'}), 400

    # 验证经纬度（只有需要位置签到时才验证）
    latitude = None
    longitude = None
    if data.get('require_checkin_location') and data.get('latitude') and data.get('longitude'):
        from utils.geolocation import validate_coordinates
        if not validate_coordinates(data['latitude'], data['longitude']):
            return jsonify({'code': 400, 'message': '经纬度格式不正确'}), 400
        latitude = data['latitude']
        longitude = data['longitude']
        print(f"[DEBUG] 创建设施 - 设施位置: 经度={longitude}, 纬度={latitude}, 半径={data.get('checkin_radius', 200)}米")
        print(f"[DEBUG] 是否需要位置签到: {data.get('require_checkin_location', True)}")

    facility = Facility(
        name=data['name'],
        category=data['category'],
        description=data.get('description', ''),
        location=data['location'],
        capacity=data.get('capacity', 1),
        image_url=data.get('image_url'),
        # 签到范围设置 - 默认无需位置验证
        latitude=latitude,
        longitude=longitude,
        checkin_radius=data.get('checkin_radius', 200),
        require_checkin_location=data.get('require_checkin_location', False)
    )

    try:
        db.session.add(facility)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': facility.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500

@admin_bp.route('/facilities/update/<int:facility_id>', methods=['PUT'])
@jwt_required()
@role_required('admin')
@log_operation('update_facility', 'facility_management')
def update_facility(facility_id):
    """更新设施信息"""
    facility = db.session.get(Facility, facility_id)

    if not facility:
        return jsonify({'code': 404, 'message': '设施不存在'}), 404

    data = request.get_json()

    # 可更新字段
    if 'name' in data:
        facility.name = data['name']
    if 'category' in data:
        facility.category = data['category']
    if 'description' in data:
        facility.description = data['description']
    if 'location' in data:
        facility.location = data['location']
    if 'capacity' in data:
        facility.capacity = data['capacity']
    if 'status' in data:
        facility.status = data['status']
    if 'image_url' in data:
        facility.image_url = data['image_url']

    # 签到范围设置
    if 'latitude' in data or 'longitude' in data:
        # 处理空字符串和null的情况
        lat_value = data.get('latitude')
        lon_value = data.get('longitude')
        
        # 空字符串转为None
        if lat_value == '':
            lat_value = None
        if lon_value == '':
            lon_value = None
        
        # 只有当两个值都有效时才更新位置信息，否则保留原值
        if lat_value is not None and lon_value is not None:
            from utils.geolocation import validate_coordinates
            if not validate_coordinates(lat_value, lon_value):
                return jsonify({'code': 400, 'message': '经纬度格式不正确'}), 400
            facility.latitude = lat_value
            facility.longitude = lon_value
            print(f"[DEBUG] 更新设施 - 设施位置: 经度={lon_value}, 纬度={lat_value}, 半径={data.get('checkin_radius', 200)}米")
        else:
            print(f"[DEBUG] 更新设施 - 保留原位置信息 (lat={facility.latitude}, lon={facility.longitude})")
    if 'checkin_radius' in data:
        facility.checkin_radius = data['checkin_radius']
        print(f"[DEBUG] 更新设施 - 签到半径: {data['checkin_radius']}米")
    if 'require_checkin_location' in data:
        facility.require_checkin_location = data['require_checkin_location']
        print(f"[DEBUG] 更新设施 - 需要位置签到: {data['require_checkin_location']}")
    if 'checkin_radius' in data:
        facility.checkin_radius = data['checkin_radius']
        print(f"[DEBUG] 更新设施 - 签到半径: {data['checkin_radius']}米")
    if 'require_checkin_location' in data:
        facility.require_checkin_location = data['require_checkin_location']
        print(f"[DEBUG] 更新设施 - 需要位置签到: {data['require_checkin_location']}")

    try:
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '更新成功',
            'data': facility.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新失败: {str(e)}'}), 500

@admin_bp.route('/facilities/delete/<int:facility_id>', methods=['DELETE'])
@jwt_required()
@role_required('admin')
@log_operation('delete_facility', 'facility_management')
def delete_facility(facility_id):
    """删除设施"""
    facility = db.session.get(Facility, facility_id)

    if not facility:
        return jsonify({'code': 404, 'message': '设施不存在'}), 404

    try:
        # 先删除所有关联记录，避免外键约束报错
        # Review、Booking、UserBehavior、Favorite、FacilityView 等表的 facility_id 不允许为 NULL
        Review.query.filter_by(facility_id=facility_id).delete()
        Booking.query.filter_by(facility_id=facility_id).delete()
        UserBehavior.query.filter_by(facility_id=facility_id).delete()
        Favorite.query.filter_by(facility_id=facility_id).delete()
        FacilityView.query.filter_by(facility_id=facility_id).delete()
        # BookingRule 的 facility_id 允许 NULL（表示全局规则），设为 NULL 再删
        BookingRule.query.filter_by(facility_id=facility_id).update({'facility_id': None})
        db.session.flush()

        db.session.delete(facility)
        db.session.commit()
        return jsonify({'code': 200, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除失败: {str(e)}'}), 500

# ==================== 预约规则管理 ====================

@admin_bp.route('/rules/list', methods=['GET'])
@jwt_required()
@role_required('admin')
def list_rules():
    """获取预约规则列表"""
    rules = BookingRule.query.all()
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': [r.to_dict() for r in rules]
    })

@admin_bp.route('/rules/create', methods=['POST'])
@jwt_required()
@role_required('admin')
@log_operation('create_rule', 'rule_management')
def create_rule():
    """创建预约规则"""
    data = request.get_json()
    
    if 'rule_name' not in data:
        return jsonify({'code': 400, 'message': '缺少规则名称'}), 400
    
    rule = BookingRule(
        rule_name=data['rule_name'],
        facility_id=data.get('facility_id'),
        max_advance_days=data.get('max_advance_days', 7),
        min_duration=data.get('min_duration', 30),
        max_duration=data.get('max_duration', 120),
        daily_limit=data.get('daily_limit', 1),
        start_time=parse_time(data.get('start_time', '08:00')),
        end_time=parse_time(data.get('end_time', '22:00'))
    )
    
    try:
        db.session.add(rule)
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': rule.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500

@admin_bp.route('/rules/update/<int:rule_id>', methods=['PUT'])
@jwt_required()
@role_required('admin')
@log_operation('update_rule', 'rule_management')
def update_rule(rule_id):
    """更新预约规则"""
    rule = db.session.get(BookingRule, rule_id)
    
    if not rule:
        return jsonify({'code': 404, 'message': '规则不存在'}), 404
    
    data = request.get_json()
    
    # 可更新字段
    if 'rule_name' in data:
        rule.rule_name = data['rule_name']
    if 'max_advance_days' in data:
        rule.max_advance_days = data['max_advance_days']
    if 'min_duration' in data:
        rule.min_duration = data['min_duration']
    if 'max_duration' in data:
        rule.max_duration = data['max_duration']
    if 'daily_limit' in data:
        rule.daily_limit = data['daily_limit']
    if 'start_time' in data:
        rule.start_time = parse_time(data['start_time'])
    if 'end_time' in data:
        rule.end_time = parse_time(data['end_time'])
    if 'status' in data:
        rule.status = data['status']
    
    try:
        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '更新成功',
            'data': rule.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新失败: {str(e)}'}), 500

@admin_bp.route('/rules/delete/<int:rule_id>', methods=['DELETE'])
@jwt_required()
@role_required('admin')
@log_operation('delete_rule', 'rule_management')
def delete_rule(rule_id):
    """删除预约规则"""
    rule = db.session.get(BookingRule, rule_id)
    
    if not rule:
        return jsonify({'code': 404, 'message': '规则不存在'}), 404
    
    try:
        db.session.delete(rule)
        db.session.commit()
        return jsonify({'code': 200, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除失败: {str(e)}'}), 500

# ==================== 数据统计 ====================

@admin_bp.route('/statistics/overview', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_statistics():
    """获取数据统计概览"""
    # 用户统计
    total_users = User.query.count()
    resident_count = User.query.filter_by(role='resident').count()
    auditor_count = User.query.filter_by(role='auditor').count()
    
    # 设施统计
    total_facilities = Facility.query.count()
    active_facilities = Facility.query.filter_by(status=1).count()
    
    # 预约统计
    total_bookings = Booking.query.count()
    pending_bookings = Booking.query.filter_by(status='pending').count()
    approved_bookings = Booking.query.filter_by(status='approved').count()
    
    # 最近7天预约趋势
    seven_days_ago = datetime.now() - timedelta(days=7)
    recent_bookings = db.session.query(
        func.date(Booking.created_at).label('date'),
        func.count(Booking.booking_id).label('count')
    ).filter(Booking.created_at >= seven_days_ago)\
     .group_by(func.date(Booking.created_at))\
     .all()
    
    # 热门设施 TOP5（按实际预约数量统计）
    facility_booking_counts = db.session.query(
        Booking.facility_id,
        func.count(Booking.booking_id).label('cnt')
    ).filter(Booking.status.in_(['pending', 'approved', 'completed']))\
     .group_by(Booking.facility_id)\
     .order_by(func.count(Booking.booking_id).desc())\
     .limit(5)\
     .all()
    
    # 获取这些设施的详细信息
    popular_facilities = []
    for row in facility_booking_counts:
        facility = db.session.get(Facility, row.facility_id)
        if facility:
            popular_facilities.append({
                'facility_id': facility.facility_id,
                'name': facility.name,
                'booking_count': row.cnt
            })
    
    # 反馈统计
    total_feedbacks = Feedback.query.count()
    pending_feedbacks = Feedback.query.filter_by(status='pending').count()
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'users': {
                'total': total_users,
                'residents': resident_count,
                'auditors': auditor_count
            },
            'facilities': {
                'total': total_facilities,
                'active': active_facilities
            },
            'bookings': {
                'total': total_bookings,
                'pending': pending_bookings,
                'approved': approved_bookings
            },
            'booking_trend': [
                {
                    'date': b.date.strftime('%Y-%m-%d'),
                    'count': b.count
                } for b in recent_bookings
            ],
            'popular_facilities': popular_facilities,
            'feedbacks': {
                'total': total_feedbacks,
                'pending': pending_feedbacks
            }
        }
    })

# ==================== 操作日志 ====================

@admin_bp.route('/logs/list', methods=['GET'])
@jwt_required()
@role_required('admin')
def list_operation_logs():
    """获取操作日志"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    
    query = OperationLog.query.order_by(OperationLog.created_at.desc())
    pagination = query.paginate(page=page, per_page=page_size, error_out=False)
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'total': pagination.total,
            'page': page,
            'page_size': page_size,
            'logs': [log.to_dict() for log in pagination.items]
        }
    })

# ==================== Excel 导出功能 ====================

@admin_bp.route('/export/excel', methods=['GET'])
@jwt_required()
@role_required('admin')
def export_excel():
    """导出 Excel 报表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    report_type = request.args.get('type', 'bookings')  # bookings, users, facilities, feedbacks, auditors
    period = request.args.get('period', 'month')  # today, week, month, year, all
    
    # 计算时间范围
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    if period == 'today':
        start_date = today_start
    elif period == 'week':
        start_date = today_start - timedelta(days=now.weekday())
    elif period == 'month':
        start_date = today_start.replace(day=1)
    elif period == 'year':
        start_date = today_start.replace(month=1, day=1)
    else:  # all
        start_date = None
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    
    # 定义样式
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, headers):
        """设置表头样式"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def auto_width(ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 根据报表类型生成数据
    if report_type == 'bookings':
        ws.title = '预约报表'
        headers = ['预约ID', '用户名', '设施名称', '预约日期', '开始时间', '结束时间', '状态', '审核时间', '审核备注']
        style_header(ws, headers)
        
        query = Booking.query
        if start_date:
            query = query.filter(Booking.created_at >= start_date)
        
        bookings = query.order_by(Booking.created_at.desc()).all()
        
        row = 2
        for booking in bookings:
            user = db.session.get(User, booking.user_id)
            facility = db.session.get(Facility, booking.facility_id)
            ws.cell(row=row, column=1, value=booking.booking_id)
            ws.cell(row=row, column=2, value=user.username if user else '未知')
            ws.cell(row=row, column=3, value=facility.name if facility else '未知')
            ws.cell(row=row, column=4, value=str(booking.booking_date))
            ws.cell(row=row, column=5, value=str(booking.start_time))
            ws.cell(row=row, column=6, value=str(booking.end_time))
            ws.cell(row=row, column=7, value=booking.status)
            ws.cell(row=row, column=8, value=str(booking.audit_time) if booking.audit_time else '')
            ws.cell(row=row, column=9, value=booking.audit_comment or '')
            row += 1
    
    elif report_type == 'users':
        ws.title = '用户报表'
        headers = ['用户ID', '用户名', '手机号', '角色', '状态', '注册时间', '预约数量']
        style_header(ws, headers)
        
        users = User.query.order_by(User.created_at.desc()).all()
        
        row = 2
        for user in users:
            # 预约数量按时间段统计
            booking_query = Booking.query.filter_by(user_id=user.user_id)
            if start_date:
                booking_query = booking_query.filter(Booking.booking_date >= start_date.date())
            booking_count = booking_query.count()
            
            ws.cell(row=row, column=1, value=user.user_id)
            ws.cell(row=row, column=2, value=user.username)
            ws.cell(row=row, column=3, value=user.phone)
            ws.cell(row=row, column=4, value='居民' if user.role == 'resident' else '审核员' if user.role == 'auditor' else '管理员')
            ws.cell(row=row, column=5, value='正常' if user.status == 1 else '禁用')
            ws.cell(row=row, column=6, value=str(user.created_at))
            ws.cell(row=row, column=7, value=booking_count)
            row += 1
    
    elif report_type == 'facilities':
        ws.title = '设施报表'
        headers = ['设施ID', '设施名称', '分类', '位置', '容量', '预约次数', '签到次数', '评分', '状态', '创建时间']
        style_header(ws, headers)
        
        facilities = Facility.query.order_by(Facility.created_at.desc()).all()
        
        # 构建时间过滤条件
        facility_filter = []
        if start_date:
            facility_filter.append(Booking.booking_date >= start_date.date())
        
        row = 2
        for facility in facilities:
            # 预约次数：已通过/已完成的预约 (approved + completed)
            booking_query = Booking.query.filter(
                Facility.facility_id == facility.facility_id,
                Booking.status.in_(['approved', 'completed'])
            )
            if start_date:
                booking_query = booking_query.filter(Booking.booking_date >= start_date.date())
            booking_count = booking_query.count()
            
            # 签到次数：已完成的预约
            checkin_query = Booking.query.filter_by(facility_id=facility.facility_id, status='completed')
            if start_date:
                checkin_query = checkin_query.filter(Booking.booking_date >= start_date.date())
            checkin_count = checkin_query.count()
            
            # 计算平均评分
            avg_rating = db.session.query(func.avg(Review.rating)).filter_by(facility_id=facility.facility_id).scalar() or 0
            
            ws.cell(row=row, column=1, value=facility.facility_id)
            ws.cell(row=row, column=2, value=facility.name)
            ws.cell(row=row, column=3, value=facility.category)
            ws.cell(row=row, column=4, value=facility.location)
            ws.cell(row=row, column=5, value=facility.capacity)
            ws.cell(row=row, column=6, value=booking_count)
            ws.cell(row=row, column=7, value=checkin_count)
            ws.cell(row=row, column=8, value=round(avg_rating, 1))
            ws.cell(row=row, column=9, value='开放' if facility.status == 1 else '关闭')
            ws.cell(row=row, column=10, value=str(facility.created_at))
            row += 1
    
    elif report_type == 'feedbacks':
        ws.title = '反馈报表'
        headers = ['反馈ID', '用户名', '类型', '内容', '状态', '回复内容', '提交时间', '回复时间']
        style_header(ws, headers)
        
        # 构建时间过滤条件
        feedback_query = Feedback.query
        if start_date:
            feedback_query = feedback_query.filter(Feedback.created_at >= start_date)
        
        feedbacks = feedback_query.order_by(Feedback.created_at.desc()).all()
        
        row = 2
        for feedback in feedbacks:
            user = db.session.get(User, feedback.user_id)
            
            type_map = {'consultation': '咨询', 'complaint': '投诉', 'suggestion': '建议', 'other': '其他'}
            status_map = {'pending': '待处理', 'replied': '已回复', 'closed': '已关闭'}
            
            ws.cell(row=row, column=1, value=feedback.feedback_id)
            ws.cell(row=row, column=2, value=user.username if user else '未知')
            ws.cell(row=row, column=3, value=type_map.get(feedback.type, feedback.type))
            ws.cell(row=row, column=4, value=feedback.content[:100] + '...' if len(feedback.content) > 100 else feedback.content)
            ws.cell(row=row, column=5, value=status_map.get(feedback.status, feedback.status))
            ws.cell(row=row, column=6, value=feedback.reply or '')
            ws.cell(row=row, column=7, value=str(feedback.created_at))
            ws.cell(row=row, column=8, value=str(feedback.reply_time) if feedback.reply_time else '')
            row += 1
    
    elif report_type == 'auditors':
        ws.title = '审核员报表'
        headers = ['审核员ID', '审核员姓名', '手机号', '总审核数', '通过数', '拒绝数', '取消数', '通过率', '注册时间']
        style_header(ws, headers)
        
        # 构建审核时间过滤条件
        auditor_filter = []
        if start_date:
            auditor_filter.append(Booking.audit_time >= start_date)
        
        auditors = User.query.filter(User.role == 'auditor').order_by(User.created_at.desc()).all()
        
        row = 2
        for auditor in auditors:
            # 总审核数：所有有审核员的预约
            base_query = Booking.query.filter(Booking.auditor_id == auditor.user_id)
            if start_date:
                base_query = base_query.filter(Booking.audit_time >= start_date)
            
            total_audited = base_query.count()
            
            # 通过数：approved + completed
            approved_query = base_query.filter(Booking.status.in_(['approved', 'completed']))
            approved = approved_query.count()
            
            # 拒绝数：rejected
            rejected = base_query.filter(Booking.status == 'rejected').count()
            
            # 取消数：cancelled
            cancelled = base_query.filter(Booking.status == 'cancelled').count()
            
            approval_rate = round(approved / total_audited * 100, 1) if total_audited > 0 else 0
            
            ws.cell(row=row, column=1, value=auditor.user_id)
            ws.cell(row=row, column=2, value=auditor.username)
            ws.cell(row=row, column=3, value=auditor.phone)
            ws.cell(row=row, column=4, value=total_audited)
            ws.cell(row=row, column=5, value=approved)
            ws.cell(row=row, column=6, value=rejected)
            ws.cell(row=row, column=7, value=cancelled)
            ws.cell(row=row, column=8, value=f'{approval_rate}%')
            ws.cell(row=row, column=9, value=str(auditor.created_at))
            row += 1
    
    else:
        return jsonify({'code': 400, 'message': '无效的报表类型'}), 400
    
    # 设置边框和行高
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).alignment = cell_alignment
        ws.row_dimensions[r].height = 25
    
    # 自动调整列宽
    auto_width(ws)
    
    # 生成文件名
    period_names = {'today': '今日', 'week': '本周', 'month': '本月', 'year': '本年', 'all': '全部'}
    type_names = {'bookings': '预约', 'users': '用户', 'facilities': '设施', 'feedbacks': '反馈', 'auditors': '审核员'}
    filename = f"{period_names.get(period, '全部')}{type_names.get(report_type, '数据')}报表_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    
    # 保存到临时文件
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回文件
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# 兼容旧版：提供统计预览数据的接口（前端使用）
@admin_bp.route('/stats/bookings', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_booking_stats():
    """获取预约统计数据（用于前端预览）"""
    period = request.args.get('period', 'all')
    
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    if period == 'today':
        start_date = today_start
    elif period == 'week':
        start_date = today_start - timedelta(days=now.weekday())
    elif period == 'month':
        start_date = today_start.replace(day=1)
    elif period == 'year':
        start_date = today_start.replace(month=1, day=1)
    else:
        start_date = None
    
    query = Booking.query
    if start_date:
        query = query.filter(Booking.created_at >= start_date)
    
    # 状态分布
    status_dist = db.session.query(
        Booking.status, func.count(Booking.booking_id)
    ).filter(Booking.booking_id.in_([b.booking_id for b in query.all()])
    ).group_by(Booking.status).all()
    
    status_distribution = {s: c for s, c in status_dist}
    
    # 每日趋势
    daily = db.session.query(
        func.date(Booking.created_at).label('date'),
        func.count(Booking.booking_id).label('count')
    ).filter(Booking.created_at >= start_date if start_date else True
    ).group_by(func.date(Booking.created_at)).all()
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'total': query.count(),
            'status_distribution': status_distribution,
            'daily': [{'date': str(d.date), 'count': d.count} for d in daily]
        }
    })

@admin_bp.route('/stats/users', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_user_stats():
    """获取用户统计数据（用于前端预览）"""
    # 活跃用户（有过预约的用户）
    active_users = db.session.query(
        User.user_id, User.username, User.phone,
        func.count(Booking.booking_id).label('booking_count')
    ).join(Booking, User.user_id == Booking.user_id
    ).group_by(User.user_id).order_by(func.count(Booking.booking_id).desc()).limit(20).all()
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'active_users': [
                {'user_id': u.user_id, 'username': u.username, 'phone': u.phone, 'booking_count': u.booking_count}
                for u in active_users
            ]
        }
    })

@admin_bp.route('/stats/facilities', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_facility_stats():
    """获取设施统计数据（用于前端预览）"""
    facilities = db.session.query(
        Facility.facility_id, Facility.name, Facility.category,
        func.count(Booking.booking_id).label('booking_count'),
        func.sum(db.case((Booking.status == 'completed', 1), else_=0)).label('checkin_count')
    ).outerjoin(Booking, Facility.facility_id == Booking.facility_id
    ).group_by(Facility.facility_id).order_by(func.count(Booking.booking_id).desc()).all()
    
    result = []
    for f in facilities:
        booking_count = f.booking_count or 0
        checkin_count = int(f.checkin_count) if f.checkin_count else 0
        usage_rate = round(checkin_count / booking_count * 100, 1) if booking_count > 0 else 0
        result.append({
            'facility_id': f.facility_id,
            'name': f.name,
            'category': f.category,
            'booking_count': booking_count,
            'checkin_count': checkin_count,
            'usage_rate': usage_rate
        })
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'rankings': result
        }
    })

@admin_bp.route('/stats/feedbacks', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_feedback_stats():
    """获取反馈统计数据（用于前端预览）"""
    status_dist = db.session.query(
        Feedback.status, func.count(Feedback.feedback_id)
    ).group_by(Feedback.status).all()
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'total': Feedback.query.count(),
            'status_distribution': {s: c for s, c in status_dist}
        }
    })

@admin_bp.route('/stats/auditors', methods=['GET'])
@jwt_required()
@role_required('admin')
def get_auditor_stats():
    """获取审核员统计数据（用于前端预览）"""
    period = request.args.get('period', 'month')
    
    # 计算时间范围
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    if period == 'today':
        start_date = today_start
    elif period == 'week':
        start_date = today_start - timedelta(days=7)
    elif period == 'month':
        start_date = today_start - timedelta(days=30)
    elif period == 'year':
        start_date = today_start - timedelta(days=365)
    else:  # all
        start_date = None
    
    auditors = User.query.filter(User.role == 'auditor').all()
    
    result = []
    for auditor in auditors:
        # 总审核数：所有有审核员的预约
        base_query = Booking.query.filter(Booking.auditor_id == auditor.user_id)
        if start_date:
            base_query = base_query.filter(Booking.created_at >= start_date)
        
        total = base_query.count()
        
        # 通过数：approved + completed
        approved = base_query.filter(Booking.status.in_(['approved', 'completed'])).count()
        
        # 拒绝数：rejected
        rejected = base_query.filter(Booking.status == 'rejected').count()
        
        # 取消数：cancelled
        cancelled = base_query.filter(Booking.status == 'cancelled').count()
        
        approval_rate = round(approved / total * 100, 1) if total > 0 else 0
        
        result.append({
            'user_id': auditor.user_id,
            'username': auditor.username,
            'phone': auditor.phone,
            'total_audited': total,
            'approved': approved,
            'rejected': rejected,
            'cancelled': cancelled,
            'approval_rate': approval_rate
        })
    
    return jsonify({
        'code': 200,
        'message': '获取成功',
        'data': {
            'auditors': result
        }
    })

